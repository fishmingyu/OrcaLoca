from openpyxl import load_workbook


def read_xlsx(path: str, worksheet_id: int = 0) -> str:
    """
    Read the xlsx file located at 'path';
    Then find the worksheet with worksheet_id (default with 0) and return its content with row index
    """
    wb = load_workbook(path)
    ws = wb.worksheets[worksheet_id]
    output = ""
    for index, row in enumerate(ws.values):
        output += f"{index+1}:{str(row)};"
    return output


def write_xlsx_cell(
    path: str, cell_id: str, value: str, value_type: str, worksheet_id: int = 0
) -> str:
    """
    Read the xlsx file located at 'path';
    Then find the worksheet with 'worksheet_id' (default with 0);
    And write the cell at 'cell_id' with 'value'.
    Supported value_type: int, float, str
    """
    if value_type == "int":
        value = int(value)
    elif value_type == "float":
        value = float(value)
    else:
        assert (
            value_type == "str"
        ), "Unsupported value_type, only [int, float, str] are supported"
    wb = load_workbook(path)
    ws = wb.worksheets[worksheet_id]
    ws[cell_id] = value
    wb.save(path)
    return "Success"
